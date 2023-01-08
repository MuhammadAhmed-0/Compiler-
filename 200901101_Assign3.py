import xml.etree.ElementTree as ET
import openpyxl

# parse the XML file
tree = ET.parse('compiler.xml')
root = tree.getroot()

# create a new Excel workbook
workbook = openpyxl.Workbook()

# get the active sheet
sheet = workbook.active

# create the headers
sheet['A1'] = 'Book_Id'
sheet['B1'] = 'Author_Name'
sheet['C1'] = 'Title'
sheet['D1'] = 'Genre'
sheet['E1'] = 'Price'
sheet['F1'] = 'Publish_date'
sheet['G1'] = 'Description'

# iterate over all the 'book' elements
for i, book in enumerate(root.findall('book')):
  # extract the book id
  book_id = book.get('id')
  sheet.cell(row=i+2, column=1).value = book_id

  # extract the author name
  author_name = book.find('author').text
  sheet.cell(row=i+2, column=2).value = author_name

  # extract the title
  title = book.find('title').text
  sheet.cell(row=i+2, column=3).value = title

  # extract the genre
  genre = book.find('genre').text
  sheet.cell(row=i+2, column=4).value = genre

  # extract the price
  price = book.find('price').text
  sheet.cell(row=i+2, column=5).value = price

  # extract the publish date
  publish_date = book.find('publish_date').text
  sheet.cell(row=i+2, column=6).value = publish_date

  # extract the description
  description = book.find('description').text
  sheet.cell(row=i+2, column=7).value = description

# save the workbook
workbook.save('data.xlsx')
